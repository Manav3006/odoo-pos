from __future__ import annotations

import base64
import csv
import json
import sqlite3
from io import BytesIO
from datetime import UTC, datetime, timedelta
from typing import Any

from flask import Blueprint, Response, current_app, g, jsonify, request

from auth import auth_required
from db import get_connection


pos_bp = Blueprint("pos", __name__, url_prefix="/api")


def _row_to_dict(row: Any) -> dict[str, Any]:
    return {key: row[key] for key in row.keys()}


def _emit_customer_event(event_name: str, payload: dict[str, Any]) -> None:
    socketio = current_app.extensions.get("socketio")
    if socketio is not None:
        socketio.emit(event_name, payload, namespace="/customer")


def _is_kitchen_display_authorized() -> bool:
    provided = (
        request.headers.get("X-Kitchen-Display-Key")
        or request.args.get("key")
        or ""
    ).strip()
    expected = (current_app.config.get("KITCHEN_DISPLAY_KEY") or "").strip()
    return bool(expected) and provided == expected


def _derive_kitchen_status(order_status: str, kitchen_status: str | None) -> str:
    if kitchen_status:
        return kitchen_status
    if order_status == "REJECTED":
        return "REJECTED"
    if order_status == "PENDING_VERIFICATION":
        return "PENDING_VERIFICATION"
    if order_status == "DRAFT":
        return "NOT_SENT"
    if order_status == "PREPARING":
        return "PREPARING"
    if order_status in {"READY", "PAID"}:
        return "COMPLETED"
    return "TO_COOK"


@pos_bp.get("/products")
@auth_required
def list_products():
    db_path = current_app.config["DB_PATH"]
    with get_connection(db_path) as connection:
        rows = connection.execute(
            """
            SELECT p.id, p.name, c.name AS category, p.price, p.unit, p.tax_rate, p.description
            FROM products p
            LEFT JOIN categories c ON c.id = p.category_id
            ORDER BY c.name, p.name;
            """
        ).fetchall()

    return jsonify([_row_to_dict(row) for row in rows])


@pos_bp.get("/floors")
@auth_required
def list_floors_with_tables():
    db_path = current_app.config["DB_PATH"]
    with get_connection(db_path) as connection:
        floors = connection.execute(
            "SELECT id, name FROM floors ORDER BY id;"
        ).fetchall()

        output = []
        for floor in floors:
            tables = connection.execute(
                """
                SELECT id, table_number, seats, is_active
                FROM tables
                WHERE floor_id = ?
                ORDER BY table_number;
                """,
                (floor["id"],),
            ).fetchall()
            output.append(
                {
                    "id": floor["id"],
                    "name": floor["name"],
                    "tables": [
                        {
                            "id": table["id"],
                            "table_number": table["table_number"],
                            "seats": table["seats"],
                            "is_active": bool(table["is_active"]),
                        }
                        for table in tables
                    ],
                }
            )

    return jsonify(output)


@pos_bp.get("/payment-methods")
@auth_required
def list_payment_methods():
    db_path = current_app.config["DB_PATH"]
    with get_connection(db_path) as connection:
        rows = connection.execute(
            """
            SELECT id, name, method_type, is_enabled, upi_id
            FROM payment_methods
            ORDER BY id;
            """
        ).fetchall()

    return jsonify(
        [
            {
                "id": row["id"],
                "name": row["name"],
                "method_type": row["method_type"],
                "is_enabled": bool(row["is_enabled"]),
                "upi_id": row["upi_id"],
            }
            for row in rows
        ]
    )


@pos_bp.patch("/payment-methods/<int:method_id>")
@auth_required
def update_payment_method(method_id: int):
    role = (g.current_user.get("role") or "").strip().lower()
    if role == "customer":
        return jsonify({"error": "Only manager/staff can update payment methods."}), 403

    payload = request.get_json(silent=True) or {}
    is_enabled_raw = payload.get("is_enabled")
    upi_id_raw = payload.get("upi_id")

    if is_enabled_raw is None and upi_id_raw is None:
        return jsonify({"error": "Nothing to update."}), 400

    db_path = current_app.config["DB_PATH"]
    with get_connection(db_path) as connection:
        existing = connection.execute(
            """
            SELECT id, name, method_type, is_enabled, upi_id
            FROM payment_methods
            WHERE id = ?
            LIMIT 1;
            """,
            (method_id,),
        ).fetchone()

        if existing is None:
            return jsonify({"error": "Payment method not found."}), 404

        next_is_enabled = (
            int(bool(is_enabled_raw))
            if is_enabled_raw is not None
            else int(existing["is_enabled"])
        )
        next_upi_id = (
            (upi_id_raw or "").strip() if upi_id_raw is not None else existing["upi_id"]
        )
        if existing["method_type"] != "UPI":
            next_upi_id = None

        connection.execute(
            """
            UPDATE payment_methods
            SET is_enabled = ?, upi_id = ?
            WHERE id = ?;
            """,
            (next_is_enabled, next_upi_id, method_id),
        )

    return jsonify(
        {
            "id": method_id,
            "name": existing["name"],
            "method_type": existing["method_type"],
            "is_enabled": bool(next_is_enabled),
            "upi_id": next_upi_id,
        }
    )


@pos_bp.post("/floors")
@auth_required
def create_floor():
    role = (g.current_user.get("role") or "").strip().lower()
    if role == "customer":
        return jsonify({"error": "Only manager/staff can create floors."}), 403

    payload = request.get_json(silent=True) or {}
    name = (payload.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Floor name is required."}), 400

    db_path = current_app.config["DB_PATH"]
    try:
        with get_connection(db_path) as connection:
            floor_id = connection.execute(
                "INSERT INTO floors (name) VALUES (?);",
                (name,),
            ).lastrowid
    except sqlite3.IntegrityError:
        return jsonify({"error": "Floor name already exists."}), 409

    return jsonify({"id": floor_id, "name": name}), 201


@pos_bp.post("/tables")
@auth_required
def create_table():
    role = (g.current_user.get("role") or "").strip().lower()
    if role == "customer":
        return jsonify({"error": "Only manager/staff can create tables."}), 403

    payload = request.get_json(silent=True) or {}
    floor_id = int(payload.get("floor_id") or 0)
    table_number = (payload.get("table_number") or "").strip()
    seats = int(payload.get("seats") or 2)

    if floor_id <= 0 or not table_number:
        return jsonify({"error": "floor_id and table_number are required."}), 400
    if seats <= 0:
        return jsonify({"error": "seats must be greater than 0."}), 400

    db_path = current_app.config["DB_PATH"]
    with get_connection(db_path) as connection:
        floor = connection.execute(
            "SELECT id FROM floors WHERE id = ? LIMIT 1;",
            (floor_id,),
        ).fetchone()
        if floor is None:
            return jsonify({"error": "Floor not found."}), 404

        try:
            table_id = connection.execute(
                """
                INSERT INTO tables (floor_id, table_number, seats, is_active)
                VALUES (?, ?, ?, 1);
                """,
                (floor_id, table_number, seats),
            ).lastrowid
        except sqlite3.IntegrityError:
            return jsonify({"error": "Table number already exists on this floor."}), 409

    return (
        jsonify(
            {
                "id": table_id,
                "floor_id": floor_id,
                "table_number": table_number,
                "seats": seats,
                "is_active": True,
            }
        ),
        201,
    )


@pos_bp.patch("/tables/<int:table_id>")
@auth_required
def update_table(table_id: int):
    role = (g.current_user.get("role") or "").strip().lower()
    if role == "customer":
        return jsonify({"error": "Only manager/staff can update tables."}), 403

    payload = request.get_json(silent=True) or {}
    is_active_raw = payload.get("is_active")
    seats_raw = payload.get("seats")

    if is_active_raw is None and seats_raw is None:
        return jsonify({"error": "Nothing to update."}), 400

    db_path = current_app.config["DB_PATH"]
    with get_connection(db_path) as connection:
        existing = connection.execute(
            """
            SELECT id, floor_id, table_number, seats, is_active
            FROM tables
            WHERE id = ?
            LIMIT 1;
            """,
            (table_id,),
        ).fetchone()
        if existing is None:
            return jsonify({"error": "Table not found."}), 404

        next_is_active = (
            int(bool(is_active_raw))
            if is_active_raw is not None
            else int(existing["is_active"])
        )
        next_seats = int(seats_raw) if seats_raw is not None else int(existing["seats"])
        if next_seats <= 0:
            return jsonify({"error": "seats must be greater than 0."}), 400

        connection.execute(
            """
            UPDATE tables
            SET is_active = ?, seats = ?
            WHERE id = ?;
            """,
            (next_is_active, next_seats, table_id),
        )

    return jsonify(
        {
            "id": table_id,
            "floor_id": existing["floor_id"],
            "table_number": existing["table_number"],
            "seats": next_seats,
            "is_active": bool(next_is_active),
        }
    )


@pos_bp.post("/products")
@auth_required
def create_product():
    role = (g.current_user.get("role") or "").strip().lower()
    if role == "customer":
        return jsonify({"error": "Only manager/staff can create products."}), 403

    payload = request.get_json(silent=True) or {}
    name = (payload.get("name") or "").strip()
    category_name = (payload.get("category") or "").strip()
    unit = (payload.get("unit") or "unit").strip() or "unit"
    description = (payload.get("description") or "").strip() or None

    try:
        price = float(payload.get("price"))
        tax_rate = float(payload.get("tax_rate") or 0)
    except (TypeError, ValueError):
        return jsonify({"error": "price and tax_rate must be numeric."}), 400

    if not name:
        return jsonify({"error": "Product name is required."}), 400
    if price < 0:
        return jsonify({"error": "price cannot be negative."}), 400
    if tax_rate < 0:
        return jsonify({"error": "tax_rate cannot be negative."}), 400

    db_path = current_app.config["DB_PATH"]
    with get_connection(db_path) as connection:
        category_id = None
        if category_name:
            category = connection.execute(
                "SELECT id FROM categories WHERE name = ? LIMIT 1;",
                (category_name,),
            ).fetchone()
            if category is None:
                category_id = connection.execute(
                    "INSERT INTO categories (name) VALUES (?);",
                    (category_name,),
                ).lastrowid
            else:
                category_id = category["id"]

        product_id = connection.execute(
            """
            INSERT INTO products (
                name,
                category_id,
                price,
                unit,
                tax_rate,
                description,
                is_active
            )
            VALUES (?, ?, ?, ?, ?, ?, 1);
            """,
            (name, category_id, price, unit, tax_rate, description),
        ).lastrowid

    return (
        jsonify(
            {
                "id": product_id,
                "name": name,
                "category": category_name or None,
                "price": round(price, 2),
                "unit": unit,
                "tax_rate": round(tax_rate, 2),
                "description": description,
                "is_active": True,
            }
        ),
        201,
    )


@pos_bp.get("/sessions/active")
@auth_required
def get_active_session():
    terminal_name = request.args.get("terminal_name", "Main Register").strip()
    db_path = current_app.config["DB_PATH"]

    with get_connection(db_path) as connection:
        terminal = connection.execute(
            "SELECT id, name FROM terminals WHERE name = ? LIMIT 1;",
            (terminal_name,),
        ).fetchone()
        if terminal is None:
            return jsonify({"active": False, "session": None})

        active = connection.execute(
            """
            SELECT id, opened_at, opening_balance, status
            FROM pos_sessions
            WHERE terminal_id = ? AND closed_at IS NULL
            LIMIT 1;
            """,
            (terminal["id"],),
        ).fetchone()

        if active is None:
            return jsonify({"active": False, "session": None})

    return jsonify(
        {
            "active": True,
            "session": {
                "session_id": active["id"],
                "terminal_name": terminal_name,
                "opened_at": active["opened_at"],
                "opening_balance": active["opening_balance"],
                "status": active["status"],
            },
        }
    )


@pos_bp.post("/sessions/open")
@auth_required
def open_session():
    payload = request.get_json(silent=True) or {}
    opening_balance = float(payload.get("opening_balance") or 0)
    terminal_name = (payload.get("terminal_name") or "Main Register").strip()
    user_id = int(g.current_user["user_id"])
    now = datetime.now(UTC).isoformat()

    db_path = current_app.config["DB_PATH"]
    with get_connection(db_path) as connection:
        terminal = connection.execute(
            """
            SELECT id
            FROM terminals
            WHERE name = ?
            LIMIT 1;
            """,
            (terminal_name,),
        ).fetchone()

        if terminal is None:
            terminal_id = connection.execute(
                "INSERT INTO terminals (name, is_active) VALUES (?, 1);",
                (terminal_name,),
            ).lastrowid
        else:
            terminal_id = terminal["id"]

        active = connection.execute(
            """
            SELECT id, opened_at
            FROM pos_sessions
            WHERE terminal_id = ? AND closed_at IS NULL
            LIMIT 1;
            """,
            (terminal_id,),
        ).fetchone()
        if active is not None:
            existing = connection.execute(
                """
                SELECT opening_balance, status
                FROM pos_sessions
                WHERE id = ?;
                """,
                (active["id"],),
            ).fetchone()
            return jsonify(
                {
                    "session_id": active["id"],
                    "terminal_name": terminal_name,
                    "opened_at": active["opened_at"],
                    "opening_balance": existing["opening_balance"] if existing else 0,
                    "status": existing["status"] if existing else "OPEN",
                    "existing_session": True,
                }
            )

        session_id = connection.execute(
            """
            INSERT INTO pos_sessions (
                terminal_id,
                opened_by,
                opened_at,
                opening_balance,
                status
            )
            VALUES (?, ?, ?, ?, 'OPEN');
            """,
            (terminal_id, user_id, now, opening_balance),
        ).lastrowid

    return (
        jsonify(
            {
                "session_id": session_id,
                "terminal_name": terminal_name,
                "opened_at": now,
                "opening_balance": opening_balance,
                "status": "OPEN",
                "existing_session": False,
            }
        ),
        201,
    )


@pos_bp.post("/sessions/<int:session_id>/close")
@auth_required
def close_session(session_id: int):
    payload = request.get_json(silent=True) or {}
    closing_balance = float(payload.get("closing_balance") or 0)
    now = datetime.now(UTC).isoformat()

    db_path = current_app.config["DB_PATH"]
    with get_connection(db_path) as connection:
        session = connection.execute(
            "SELECT id, status FROM pos_sessions WHERE id = ?;",
            (session_id,),
        ).fetchone()
        if session is None:
            return jsonify({"error": "Session not found."}), 404
        if session["status"] != "OPEN":
            return jsonify({"error": "Session is already closed."}), 409

        totals = connection.execute(
            """
            SELECT COALESCE(SUM(total_amount), 0) AS total_sales
            FROM orders
            WHERE session_id = ? AND order_status = 'PAID';
            """,
            (session_id,),
        ).fetchone()

        connection.execute(
            """
            UPDATE pos_sessions
            SET closed_at = ?,
                closing_balance = ?,
                closing_sales = ?,
                status = 'CLOSED'
            WHERE id = ?;
            """,
            (now, closing_balance, totals["total_sales"], session_id),
        )

    return jsonify(
        {
            "session_id": session_id,
            "closed_at": now,
            "closing_balance": closing_balance,
            "closing_sales": totals["total_sales"],
        }
    )


@pos_bp.post("/orders")
@auth_required
def create_order():
    payload = request.get_json(silent=True) or {}
    if not payload and request.data:
        try:
            payload = json.loads(request.data.decode("utf-8"))
        except (ValueError, UnicodeDecodeError):
            payload = {}
    session_id = payload.get("session_id")
    table_id = payload.get("table_id")
    items = payload.get("items") or []
    user_role = (g.current_user.get("role") or "staff").strip().lower()
    current_user_id = int(g.current_user["user_id"])
    requested_source = (payload.get("source") or "").strip().upper()

    if requested_source == "SELF_ORDER" and user_role != "customer":
        return jsonify({"error": "Only customers can place self orders."}), 403

    is_self_order = user_role == "customer"
    initial_status = "PENDING_VERIFICATION" if is_self_order else "DRAFT"
    customer_id = current_user_id if user_role == "customer" else None

    if not session_id or not table_id or not items:
        return jsonify({"error": "session_id, table_id and items are required."}), 400

    db_path = current_app.config["DB_PATH"]
    now = datetime.now(UTC).isoformat()

    with get_connection(db_path) as connection:
        session = connection.execute(
            "SELECT id FROM pos_sessions WHERE id = ? AND status = 'OPEN';",
            (session_id,),
        ).fetchone()
        if session is None:
            return jsonify({"error": "Active session not found."}), 400

        table = connection.execute(
            "SELECT id, is_active FROM tables WHERE id = ?;",
            (table_id,),
        ).fetchone()
        if table is None or not table["is_active"]:
            return jsonify({"error": "Table not available."}), 400

        day_stamp = datetime.now(UTC).strftime("%Y%m%d")
        order_count = connection.execute(
            "SELECT COUNT(*) AS count_for_day FROM orders WHERE order_number LIKE ?;",
            (f"ORD-{day_stamp}-%",),
        ).fetchone()["count_for_day"]
        order_number = f"ORD-{day_stamp}-{order_count + 1:04d}"

        subtotal = 0.0
        total_tax = 0.0
        expanded_items: list[dict[str, Any]] = []
        for item in items:
            product_id = int(item.get("product_id") or 0)
            quantity = int(item.get("quantity") or 00)
            if product_id <= 0 or quantity <= 0:
                return jsonify({"error": "Invalid product_id or quantity."}), 400

            product = connection.execute(
                """
                SELECT id, name, price, tax_rate
                FROM products
                WHERE id = ?;
                """,
                (product_id,),
            ).fetchone()
            if product is None:
                return jsonify({"error": f"Product {product_id} not found."}), 404

            line_subtotal = float(product["price"]) * quantity
            line_tax = line_subtotal * (float(product["tax_rate"]) / 100.0)
            subtotal += line_subtotal
            total_tax += line_tax
            expanded_items.append(
                {
                    "product_id": product["id"],
                    "product_name": product["name"],
                    "quantity": quantity,
                    "unit_price": float(product["price"]),
                    "line_subtotal": round(line_subtotal, 2),
                    "line_tax": round(line_tax, 2),
                }
            )

        total_amount = round(subtotal + total_tax, 2)

        order_id = connection.execute(
            """
            INSERT INTO orders (
                order_number,
                session_id,
                table_id,
                customer_id,
                created_at,
                order_status,
                subtotal,
                tax_total,
                total_amount
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?);
            """,
            (
                order_number,
                session_id,
                table_id,
                customer_id,
                now,
                initial_status,
                subtotal,
                total_tax,
                total_amount,
            ),
        ).lastrowid

        connection.execute(
            """
            INSERT INTO order_status_history (
                order_id,
                previous_status,
                next_status,
                changed_by,
                changed_at
            )
            VALUES (?, NULL, ?, ?, ?);
            """,
            (order_id, initial_status, current_user_id, now),
        )

        for item in expanded_items:
            connection.execute(
                """
                INSERT INTO order_items (
                    order_id,
                    product_id,
                    quantity,
                    unit_price,
                    line_subtotal,
                    line_tax
                )
                VALUES (?, ?, ?, ?, ?, ?);
                """,
                (
                    order_id,
                    item["product_id"],
                    item["quantity"],
                    item["unit_price"],
                    item["line_subtotal"],
                    item["line_tax"],
                ),
            )

    if initial_status == "PENDING_VERIFICATION":
        _emit_customer_event(
            "customer:order_pending_verification",
            {
                "order_id": order_id,
                "order_number": order_number,
                "order_status": initial_status,
                "created_at": now,
            },
        )

    return (
        jsonify(
            {
                "order_id": order_id,
                "order_number": order_number,
                "order_status": initial_status,
                "requires_manager_verification": initial_status == "PENDING_VERIFICATION",
                "customer_id": customer_id,
                "subtotal": round(subtotal, 2),
                "tax_total": round(total_tax, 2),
                "total_amount": total_amount,
                "items": expanded_items,
            }
        ),
        201,
    )


@pos_bp.post("/orders/<int:order_id>/send-kitchen")
@auth_required
def send_order_to_kitchen(order_id: int):
    now = datetime.now(UTC).isoformat()
    db_path = current_app.config["DB_PATH"]
    changed_by = int(g.current_user["user_id"])
    requester_role = (g.current_user.get("role") or "").strip().lower()

    if requester_role == "customer":
        return jsonify({"error": "Customer orders must be verified by manager/staff."}), 403

    with get_connection(db_path) as connection:
        order = connection.execute(
            """
            SELECT id, order_number, table_id, order_status
            FROM orders
            WHERE id = ?;
            """,
            (order_id,),
        ).fetchone()
        if order is None:
            return jsonify({"error": "Order not found."}), 404

        if order["order_status"] == "SENT_TO_KITCHEN":
            return jsonify({"error": "Order is already in kitchen queue."}), 409

        if order["order_status"] not in {"DRAFT", "PENDING_VERIFICATION"}:
            return jsonify({"error": "Order cannot be sent to kitchen."}), 409

        existing_ticket = connection.execute(
            """
            SELECT id
            FROM kitchen_tickets
            WHERE order_id = ?
            ORDER BY id DESC
            LIMIT 1;
            """,
            (order_id,),
        ).fetchone()
        if existing_ticket is not None:
            return jsonify({"error": "Order already has a kitchen ticket."}), 409

        table = connection.execute(
            "SELECT table_number FROM tables WHERE id = ?;",
            (order["table_id"],),
        ).fetchone()

        line_items = connection.execute(
            """
            SELECT oi.product_id, p.name, oi.quantity
            FROM order_items oi
            JOIN products p ON p.id = oi.product_id
            WHERE oi.order_id = ?;
            """,
            (order_id,),
        ).fetchall()

        ticket_id = connection.execute(
            """
            INSERT INTO kitchen_tickets (
                order_id,
                ticket_status,
                created_at,
                updated_at
            )
            VALUES (?, 'TO_COOK', ?, ?);
            """,
            (order_id, now, now),
        ).lastrowid

        connection.execute(
            "UPDATE orders SET order_status = 'SENT_TO_KITCHEN' WHERE id = ?;",
            (order_id,),
        )

        connection.execute(
            """
            INSERT INTO order_status_history (
                order_id,
                previous_status,
                next_status,
                changed_by,
                changed_at
            )
            VALUES (?, ?, 'SENT_TO_KITCHEN', ?, ?);
            """,
            (order_id, order["order_status"], changed_by, now),
        )

    event_payload = {
        "ticket_id": ticket_id,
        "order_id": order_id,
        "order_number": order["order_number"],
        "table_number": table["table_number"] if table else "-",
        "ticket_status": "TO_COOK",
        "items": [
            {
                "product_id": item["product_id"],
                "product_name": item["name"],
                "quantity": item["quantity"],
            }
            for item in line_items
        ],
        "created_at": now,
    }

    socketio = current_app.extensions.get("socketio")
    if socketio is not None:
        socketio.emit("kitchen:ticket_created", event_payload, namespace="/kitchen")
    _emit_customer_event(
        "customer:order_sent",
        {
            "order_id": order_id,
            "order_number": order["order_number"],
            "table_number": table["table_number"] if table else "-",
            "kitchen_status": "TO_COOK",
            "sent_at": now,
        },
    )

    return jsonify(event_payload), 201


@pos_bp.get("/orders/pending-verification")
@auth_required
def list_pending_verification_orders():
    role = (g.current_user.get("role") or "").strip().lower()
    if role == "customer":
        return jsonify({"error": "Only manager/staff can verify orders."}), 403

    session_id_raw = (request.args.get("session_id") or "").strip()
    params: list[Any] = []
    where_clauses = ["o.order_status = 'PENDING_VERIFICATION'"]

    if session_id_raw:
        try:
            session_id = int(session_id_raw)
        except ValueError:
            return jsonify({"error": "session_id must be an integer."}), 400
        where_clauses.append("o.session_id = ?")
        params.append(session_id)

    db_path = current_app.config["DB_PATH"]
    where_sql = " AND ".join(where_clauses)

    with get_connection(db_path) as connection:
        orders = connection.execute(
            f"""
            SELECT
                o.id,
                o.order_number,
                o.session_id,
                o.table_id,
                o.customer_id,
                o.total_amount,
                o.created_at,
                t.table_number,
                cu.username AS customer_username,
                cu.email AS customer_email
            FROM orders o
            JOIN tables t ON t.id = o.table_id
            LEFT JOIN users cu ON cu.id = o.customer_id
            WHERE {where_sql}
            ORDER BY o.id ASC;
            """,
            tuple(params),
        ).fetchall()

        output = []
        for order in orders:
            items = connection.execute(
                """
                SELECT p.name AS product_name, oi.quantity
                FROM order_items oi
                JOIN products p ON p.id = oi.product_id
                WHERE oi.order_id = ?
                ORDER BY oi.id;
                """,
                (order["id"],),
            ).fetchall()

            output.append(
                {
                    "order_id": order["id"],
                    "order_number": order["order_number"],
                    "session_id": order["session_id"],
                    "table_id": order["table_id"],
                    "table_number": order["table_number"],
                    "customer_id": order["customer_id"],
                    "customer_username": order["customer_username"],
                    "customer_email": order["customer_email"],
                    "total_amount": order["total_amount"],
                    "created_at": order["created_at"],
                    "items": [
                        {
                            "product_name": item["product_name"],
                            "quantity": item["quantity"],
                        }
                        for item in items
                    ],
                }
            )

    return jsonify(output)


@pos_bp.patch("/orders/<int:order_id>/reject")
@auth_required
def reject_pending_order(order_id: int):
    role = (g.current_user.get("role") or "").strip().lower()
    if role == "customer":
        return jsonify({"error": "Only manager/staff can reject orders."}), 403

    payload = request.get_json(silent=True) or {}
    reason = (payload.get("reason") or "Rejected by manager.").strip()
    now = datetime.now(UTC).isoformat()
    changed_by = int(g.current_user["user_id"])
    db_path = current_app.config["DB_PATH"]

    with get_connection(db_path) as connection:
        order = connection.execute(
            """
            SELECT id, order_number, order_status, table_id
            FROM orders
            WHERE id = ?
            LIMIT 1;
            """,
            (order_id,),
        ).fetchone()

        if order is None:
            return jsonify({"error": "Order not found."}), 404

        if order["order_status"] != "PENDING_VERIFICATION":
            return jsonify({"error": "Only pending verification orders can be rejected."}), 409

        table = connection.execute(
            "SELECT table_number FROM tables WHERE id = ? LIMIT 1;",
            (order["table_id"],),
        ).fetchone()

        connection.execute(
            "UPDATE orders SET order_status = 'REJECTED' WHERE id = ?;",
            (order_id,),
        )

        connection.execute(
            """
            INSERT INTO order_status_history (
                order_id,
                previous_status,
                next_status,
                changed_by,
                changed_at
            )
            VALUES (?, ?, 'REJECTED', ?, ?);
            """,
            (order_id, order["order_status"], changed_by, now),
        )

    _emit_customer_event(
        "customer:order_rejected",
        {
            "order_id": order_id,
            "order_number": order["order_number"],
            "table_number": table["table_number"] if table else "-",
            "order_status": "REJECTED",
            "reason": reason,
            "updated_at": now,
        },
    )

    return jsonify(
        {
            "order_id": order_id,
            "order_number": order["order_number"],
            "order_status": "REJECTED",
            "reason": reason,
            "updated_at": now,
        }
    )


@pos_bp.post("/orders/<int:order_id>/payments/upi-qr")
@auth_required
def generate_upi_qr(order_id: int):
    try:
        import qrcode
    except ModuleNotFoundError:
        return (
            jsonify(
                {
                    "error": "UPI QR dependency missing. Install with: pip install qrcode Pillow"
                }
            ),
            500,
        )

    now = datetime.now(UTC).isoformat()
    db_path = current_app.config["DB_PATH"]

    with get_connection(db_path) as connection:
        order = connection.execute(
            """
            SELECT id, order_number, order_status, total_amount
            FROM orders
            WHERE id = ?;
            """,
            (order_id,),
        ).fetchone()
        if order is None:
            return jsonify({"error": "Order not found."}), 404
        if order["order_status"] == "PAID":
            return jsonify({"error": "Order is already paid."}), 409
        if order["order_status"] in {"PENDING_VERIFICATION", "REJECTED"}:
            return jsonify({"error": "Order cannot be paid before manager approval."}), 409

        method = connection.execute(
            """
            SELECT id, name, upi_id
            FROM payment_methods
            WHERE method_type = 'UPI' AND is_enabled = 1
            LIMIT 1;
            """
        ).fetchone()
        if method is None or not method["upi_id"]:
            return jsonify({"error": "UPI payment method is not configured."}), 400

        amount = round(float(order["total_amount"]), 2)
        upi_uri = (
            f"upi://pay?pa={method['upi_id']}&pn=Odoo%20POS%20Cafe"
            f"&am={amount:.2f}&cu=INR&tn={order['order_number']}"
        )

        qr = qrcode.QRCode(version=1, box_size=6, border=2)
        qr.add_data(upi_uri)
        qr.make(fit=True)
        image = qr.make_image(fill_color="#17310f", back_color="white")
        buffer = BytesIO()
        image.save(buffer, format="PNG")
        qr_image_data_url = (
            "data:image/png;base64,"
            + base64.b64encode(buffer.getvalue()).decode("ascii")
        )

        payment_id = connection.execute(
            """
            INSERT INTO payments (
                order_id,
                payment_method_id,
                amount,
                payment_status,
                reference_code,
                created_at
            )
            VALUES (?, ?, ?, 'PENDING', ?, ?);
            """,
            (order_id, method["id"], amount, f"UPI_PENDING_{order['order_number']}", now),
        ).lastrowid

    return jsonify(
        {
            "payment_id": payment_id,
            "order_id": order_id,
            "order_number": order["order_number"],
            "amount": amount,
            "method": {
                "id": method["id"],
                "name": method["name"],
                "upi_id": method["upi_id"],
            },
            "upi_uri": upi_uri,
            "qr_image_data_url": qr_image_data_url,
        }
    )


@pos_bp.post("/orders/<int:order_id>/payments/confirm")
@auth_required
def confirm_payment(order_id: int):
    payload = request.get_json(silent=True) or {}
    payment_id = payload.get("payment_id")
    payment_method_id = payload.get("payment_method_id")
    reference_code = (payload.get("reference_code") or "").strip()
    now = datetime.now(UTC).isoformat()
    db_path = current_app.config["DB_PATH"]
    changed_by = int(g.current_user["user_id"])

    with get_connection(db_path) as connection:
        order = connection.execute(
            """
            SELECT id, order_number, order_status, total_amount
            FROM orders
            WHERE id = ?;
            """,
            (order_id,),
        ).fetchone()
        if order is None:
            return jsonify({"error": "Order not found."}), 404
        if order["order_status"] == "PAID":
            return jsonify({"error": "Order is already paid."}), 409
        if order["order_status"] in {"PENDING_VERIFICATION", "REJECTED"}:
            return jsonify({"error": "Order cannot be paid before manager approval."}), 409

        selected_payment = None
        selected_method = None

        if payment_id is not None:
            selected_payment = connection.execute(
                """
                SELECT id, payment_method_id, amount, payment_status
                FROM payments
                WHERE id = ? AND order_id = ?;
                """,
                (payment_id, order_id),
            ).fetchone()
            if selected_payment is None:
                return jsonify({"error": "Payment record not found for this order."}), 404

            selected_method = connection.execute(
                """
                SELECT id, name, method_type, is_enabled
                FROM payment_methods
                WHERE id = ?;
                """,
                (selected_payment["payment_method_id"],),
            ).fetchone()
            if selected_method is None or not selected_method["is_enabled"]:
                return jsonify({"error": "Payment method is disabled."}), 409

            connection.execute(
                """
                UPDATE payments
                SET payment_status = 'CONFIRMED',
                    reference_code = ?,
                    created_at = ?
                WHERE id = ?;
                """,
                (
                    reference_code or f"CONFIRMED_{order['order_number']}",
                    now,
                    selected_payment["id"],
                ),
            )
            paid_amount = round(float(selected_payment["amount"]), 2)
        else:
            if payment_method_id is None:
                return jsonify({"error": "payment_method_id or payment_id is required."}), 400

            selected_method = connection.execute(
                """
                SELECT id, name, method_type, is_enabled
                FROM payment_methods
                WHERE id = ?;
                """,
                (payment_method_id,),
            ).fetchone()
            if selected_method is None:
                return jsonify({"error": "Payment method not found."}), 404
            if not selected_method["is_enabled"]:
                return jsonify({"error": "Payment method is disabled."}), 409

            paid_amount = round(float(order["total_amount"]), 2)
            payment_id = connection.execute(
                """
                INSERT INTO payments (
                    order_id,
                    payment_method_id,
                    amount,
                    payment_status,
                    reference_code,
                    created_at
                )
                VALUES (?, ?, ?, 'CONFIRMED', ?, ?);
                """,
                (
                    order_id,
                    selected_method["id"],
                    paid_amount,
                    reference_code or f"CONFIRMED_{order['order_number']}",
                    now,
                ),
            ).lastrowid

        connection.execute(
            "UPDATE orders SET order_status = 'PAID' WHERE id = ?;",
            (order_id,),
        )
        connection.execute(
            """
            INSERT INTO order_status_history (
                order_id,
                previous_status,
                next_status,
                changed_by,
                changed_at
            )
            VALUES (?, ?, 'PAID', ?, ?);
            """,
            (order_id, order["order_status"], changed_by, now),
        )

    event_payload = {
        "order_id": order_id,
        "order_number": order["order_number"],
        "payment_status": "PAID",
        "amount": paid_amount,
        "paid_at": now,
    }
    _emit_customer_event("customer:payment_confirmed", event_payload)

    return jsonify(
        {
            "order_id": order_id,
            "order_number": order["order_number"],
            "order_status": "PAID",
            "payment": {
                "payment_id": payment_id,
                "payment_method_id": selected_method["id"],
                "payment_method": selected_method["name"],
                "payment_type": selected_method["method_type"],
                "amount": paid_amount,
                "reference_code": reference_code or f"CONFIRMED_{order['order_number']}",
            },
        }
    )


@pos_bp.get("/kitchen/tickets")
@auth_required
def list_kitchen_tickets():
    status_filter = (request.args.get("status") or "").strip().upper()
    db_path = current_app.config["DB_PATH"]

    query = """
        SELECT
            kt.id,
            kt.order_id,
            kt.ticket_status,
            kt.created_at,
            kt.updated_at,
            o.order_number,
            t.table_number
        FROM kitchen_tickets kt
        JOIN orders o ON o.id = kt.order_id
        LEFT JOIN tables t ON t.id = o.table_id
    """
    params: tuple[Any, ...] = ()
    if status_filter:
        query += " WHERE kt.ticket_status = ?"
        params = (status_filter,)
    query += " ORDER BY kt.created_at DESC;"

    with get_connection(db_path) as connection:
        tickets = connection.execute(query, params).fetchall()
        output = []
        for ticket in tickets:
            items = connection.execute(
                """
                SELECT p.name AS product_name, oi.quantity
                FROM order_items oi
                JOIN products p ON p.id = oi.product_id
                WHERE oi.order_id = ?;
                """,
                (ticket["order_id"],),
            ).fetchall()
            output.append(
                {
                    "ticket_id": ticket["id"],
                    "order_id": ticket["order_id"],
                    "order_number": ticket["order_number"],
                    "table_number": ticket["table_number"],
                    "ticket_status": ticket["ticket_status"],
                    "created_at": ticket["created_at"],
                    "updated_at": ticket["updated_at"],
                    "items": [
                        {"product_name": item["product_name"], "quantity": item["quantity"]}
                        for item in items
                    ],
                }
            )

    return jsonify(output)


@pos_bp.get("/kitchen/public/tickets")
def list_public_kitchen_tickets():
    if not _is_kitchen_display_authorized():
        return jsonify({"error": "Invalid kitchen display key."}), 403

    status_filter = (request.args.get("status") or "").strip().upper()
    db_path = current_app.config["DB_PATH"]

    query = """
        SELECT
            kt.id,
            kt.order_id,
            kt.ticket_status,
            kt.created_at,
            kt.updated_at,
            o.order_number,
            t.table_number
        FROM kitchen_tickets kt
        JOIN orders o ON o.id = kt.order_id
        LEFT JOIN tables t ON t.id = o.table_id
    """
    params: tuple[Any, ...] = ()
    if status_filter:
        query += " WHERE kt.ticket_status = ?"
        params = (status_filter,)
    query += " ORDER BY kt.created_at DESC;"

    with get_connection(db_path) as connection:
        tickets = connection.execute(query, params).fetchall()
        output = []
        for ticket in tickets:
            items = connection.execute(
                """
                SELECT p.name AS product_name, oi.quantity
                FROM order_items oi
                JOIN products p ON p.id = oi.product_id
                WHERE oi.order_id = ?;
                """,
                (ticket["order_id"],),
            ).fetchall()
            output.append(
                {
                    "ticket_id": ticket["id"],
                    "order_id": ticket["order_id"],
                    "order_number": ticket["order_number"],
                    "table_number": ticket["table_number"],
                    "ticket_status": ticket["ticket_status"],
                    "created_at": ticket["created_at"],
                    "updated_at": ticket["updated_at"],
                    "items": [
                        {"product_name": item["product_name"], "quantity": item["quantity"]}
                        for item in items
                    ],
                }
            )

    return jsonify(output)


@pos_bp.patch("/kitchen/tickets/<int:ticket_id>/status")
@auth_required
def update_kitchen_ticket_status(ticket_id: int):
    payload = request.get_json(silent=True) or {}
    requested_status = (payload.get("status") or "").strip().upper()
    valid_statuses = ["TO_COOK", "PREPARING", "COMPLETED"]
    if requested_status not in valid_statuses:
        return jsonify({"error": "Invalid kitchen status."}), 400

    transitions = {
        "TO_COOK": ["PREPARING"],
        "PREPARING": ["COMPLETED"],
        "COMPLETED": [],
    }

    now = datetime.now(UTC).isoformat()
    db_path = current_app.config["DB_PATH"]
    changed_by = int(g.current_user["user_id"])

    with get_connection(db_path) as connection:
        ticket = connection.execute(
            """
            SELECT id, order_id, ticket_status
            FROM kitchen_tickets
            WHERE id = ?;
            """,
            (ticket_id,),
        ).fetchone()
        if ticket is None:
            return jsonify({"error": "Kitchen ticket not found."}), 404

        current_status = ticket["ticket_status"]
        if requested_status == current_status:
            return jsonify(
                {
                    "ticket_id": ticket_id,
                    "order_id": ticket["order_id"],
                    "ticket_status": current_status,
                    "updated_at": now,
                }
            )

        if requested_status not in transitions.get(current_status, []):
            return (
                jsonify(
                    {
                        "error": f"Invalid transition from {current_status} to {requested_status}."
                    }
                ),
                409,
            )

        connection.execute(
            """
            UPDATE kitchen_tickets
            SET ticket_status = ?, updated_at = ?
            WHERE id = ?;
            """,
            (requested_status, now, ticket_id),
        )

        order_status = "SENT_TO_KITCHEN"
        if requested_status == "PREPARING":
            order_status = "PREPARING"
        if requested_status == "COMPLETED":
            order_status = "READY"

        previous_order_status = connection.execute(
            "SELECT order_status FROM orders WHERE id = ?;",
            (ticket["order_id"],),
        ).fetchone()

        connection.execute(
            "UPDATE orders SET order_status = ? WHERE id = ?;",
            (order_status, ticket["order_id"]),
        )
        connection.execute(
            """
            INSERT INTO order_status_history (
                order_id,
                previous_status,
                next_status,
                changed_by,
                changed_at
            )
            VALUES (?, ?, ?, ?, ?);
            """,
            (
                ticket["order_id"],
                previous_order_status["order_status"] if previous_order_status else None,
                order_status,
                changed_by,
                now,
            ),
        )

    payload_out = {
        "ticket_id": ticket_id,
        "order_id": ticket["order_id"],
        "ticket_status": requested_status,
        "updated_at": now,
    }

    socketio = current_app.extensions.get("socketio")
    if socketio is not None:
        socketio.emit("kitchen:ticket_updated", payload_out, namespace="/kitchen")

    _emit_customer_event(
        "customer:kitchen_status_updated",
        {
            "order_id": ticket["order_id"],
            "kitchen_status": requested_status,
            "updated_at": now,
        },
    )

    return jsonify(payload_out)


@pos_bp.patch("/kitchen/public/tickets/<int:ticket_id>/status")
def update_public_kitchen_ticket_status(ticket_id: int):
    if not _is_kitchen_display_authorized():
        return jsonify({"error": "Invalid kitchen display key."}), 403

    payload = request.get_json(silent=True) or {}
    requested_status = (payload.get("status") or "").strip().upper()
    valid_statuses = ["TO_COOK", "PREPARING", "COMPLETED"]
    if requested_status not in valid_statuses:
        return jsonify({"error": "Invalid kitchen status."}), 400

    transitions = {
        "TO_COOK": ["PREPARING"],
        "PREPARING": ["COMPLETED"],
        "COMPLETED": [],
    }

    now = datetime.now(UTC).isoformat()
    db_path = current_app.config["DB_PATH"]

    with get_connection(db_path) as connection:
        ticket = connection.execute(
            """
            SELECT id, order_id, ticket_status
            FROM kitchen_tickets
            WHERE id = ?;
            """,
            (ticket_id,),
        ).fetchone()
        if ticket is None:
            return jsonify({"error": "Kitchen ticket not found."}), 404

        current_status = ticket["ticket_status"]
        if requested_status == current_status:
            return jsonify(
                {
                    "ticket_id": ticket_id,
                    "order_id": ticket["order_id"],
                    "ticket_status": current_status,
                    "updated_at": now,
                }
            )

        if requested_status not in transitions.get(current_status, []):
            return (
                jsonify(
                    {
                        "error": f"Invalid transition from {current_status} to {requested_status}."
                    }
                ),
                409,
            )

        connection.execute(
            """
            UPDATE kitchen_tickets
            SET ticket_status = ?, updated_at = ?
            WHERE id = ?;
            """,
            (requested_status, now, ticket_id),
        )

        order_status = "SENT_TO_KITCHEN"
        if requested_status == "PREPARING":
            order_status = "PREPARING"
        if requested_status == "COMPLETED":
            order_status = "READY"

        previous_order_status = connection.execute(
            "SELECT order_status FROM orders WHERE id = ?;",
            (ticket["order_id"],),
        ).fetchone()

        connection.execute(
            "UPDATE orders SET order_status = ? WHERE id = ?;",
            (order_status, ticket["order_id"]),
        )
        connection.execute(
            """
            INSERT INTO order_status_history (
                order_id,
                previous_status,
                next_status,
                changed_by,
                changed_at
            )
            VALUES (?, ?, ?, NULL, ?);
            """,
            (
                ticket["order_id"],
                previous_order_status["order_status"] if previous_order_status else None,
                order_status,
                now,
            ),
        )

    payload_out = {
        "ticket_id": ticket_id,
        "order_id": ticket["order_id"],
        "ticket_status": requested_status,
        "updated_at": now,
    }

    socketio = current_app.extensions.get("socketio")
    if socketio is not None:
        socketio.emit("kitchen:ticket_updated", payload_out, namespace="/kitchen")

    _emit_customer_event(
        "customer:kitchen_status_updated",
        {
            "order_id": ticket["order_id"],
            "kitchen_status": requested_status,
            "updated_at": now,
        },
    )

    return jsonify(payload_out)


@pos_bp.get("/customer-display/latest")
@auth_required
def get_customer_display_latest():
    table_id_param = (request.args.get("table_id") or "").strip()
    db_path = current_app.config["DB_PATH"]
    role = (g.current_user.get("role") or "").strip().lower()
    current_user_id = int(g.current_user["user_id"])

    query = """
        SELECT
            o.id,
            o.order_number,
            o.order_status,
            o.total_amount,
            o.created_at,
            o.table_id,
            t.table_number,
            (
                SELECT p.payment_status
                FROM payments p
                WHERE p.order_id = o.id
                ORDER BY p.id DESC
                LIMIT 1
            ) AS payment_status,
            (
                SELECT kt.ticket_status
                FROM kitchen_tickets kt
                WHERE kt.order_id = o.id
                ORDER BY kt.id DESC
                LIMIT 1
            ) AS kitchen_status
        FROM orders o
        LEFT JOIN tables t ON t.id = o.table_id
    """

    where_clauses: list[str] = []
    params_list: list[Any] = []
    if role == "customer":
        where_clauses.append("o.customer_id = ?")
        params_list.append(current_user_id)

    if table_id_param:
        try:
            table_id = int(table_id_param)
        except ValueError:
            return jsonify({"error": "table_id must be an integer."}), 400

        where_clauses.append("o.table_id = ?")
        params_list.append(table_id)

    if where_clauses:
        query += " WHERE " + " AND ".join(where_clauses)

    query += " ORDER BY o.id DESC LIMIT 1;"
    params = tuple(params_list)

    with get_connection(db_path) as connection:
        order = connection.execute(query, params).fetchone()
        if order is None:
            return jsonify({"available": False, "order": None})

        items = connection.execute(
            """
            SELECT p.name AS product_name, oi.quantity, oi.line_subtotal
            FROM order_items oi
            JOIN products p ON p.id = oi.product_id
            WHERE oi.order_id = ?
            ORDER BY oi.id;
            """,
            (order["id"],),
        ).fetchall()

    payment_state = order["payment_status"] or "UNPAID"
    if order["order_status"] == "PAID":
        payment_state = "PAID"

    return jsonify(
        {
            "available": True,
            "order": {
                "order_id": order["id"],
                "order_number": order["order_number"],
                "order_status": order["order_status"],
                "kitchen_status": _derive_kitchen_status(
                    order["order_status"], order["kitchen_status"]
                ),
                "payment_status": payment_state,
                "table_id": order["table_id"],
                "table_number": order["table_number"],
                "total_amount": order["total_amount"],
                "created_at": order["created_at"],
                "items": [
                    {
                        "product_name": item["product_name"],
                        "quantity": item["quantity"],
                        "line_subtotal": item["line_subtotal"],
                    }
                    for item in items
                ],
            },
        }
    )


@pos_bp.get("/customer-display/order/<int:order_id>")
@auth_required
def get_customer_order_status(order_id: int):
    db_path = current_app.config["DB_PATH"]
    role = (g.current_user.get("role") or "").strip().lower()
    current_user_id = int(g.current_user["user_id"])

    where_clause = "o.id = ?"
    params: tuple[Any, ...] = (order_id,)
    if role == "customer":
        where_clause += " AND o.customer_id = ?"
        params = (order_id, current_user_id)

    query = f"""
        SELECT
            o.id,
            o.order_number,
            o.order_status,
            o.total_amount,
            o.created_at,
            o.table_id,
            t.table_number,
            (
                SELECT p.payment_status
                FROM payments p
                WHERE p.order_id = o.id
                ORDER BY p.id DESC
                LIMIT 1
            ) AS payment_status,
            (
                SELECT kt.ticket_status
                FROM kitchen_tickets kt
                WHERE kt.order_id = o.id
                ORDER BY kt.id DESC
                LIMIT 1
            ) AS kitchen_status
        FROM orders o
        LEFT JOIN tables t ON t.id = o.table_id
        WHERE {where_clause}
        LIMIT 1;
    """

    with get_connection(db_path) as connection:
        order = connection.execute(
            query,
            params,
        ).fetchone()

        if order is None:
            return jsonify({"error": "Order not found."}), 404

        items = connection.execute(
            """
            SELECT p.name AS product_name, oi.quantity, oi.line_subtotal
            FROM order_items oi
            JOIN products p ON p.id = oi.product_id
            WHERE oi.order_id = ?
            ORDER BY oi.id;
            """,
            (order_id,),
        ).fetchall()

    payment_state = order["payment_status"] or "UNPAID"
    if order["order_status"] == "PAID":
        payment_state = "PAID"

    return jsonify(
        {
            "available": True,
            "order": {
                "order_id": order["id"],
                "order_number": order["order_number"],
                "order_status": order["order_status"],
                "kitchen_status": _derive_kitchen_status(
                    order["order_status"], order["kitchen_status"]
                ),
                "payment_status": payment_state,
                "table_id": order["table_id"],
                "table_number": order["table_number"],
                "total_amount": order["total_amount"],
                "created_at": order["created_at"],
                "items": [
                    {
                        "product_name": item["product_name"],
                        "quantity": item["quantity"],
                        "line_subtotal": item["line_subtotal"],
                    }
                    for item in items
                ],
            },
        }
    )


@pos_bp.get("/reports/sales")
@auth_required
def get_sales_report():
    period = (request.args.get("period") or "all").strip().lower()
    session_id_raw = (request.args.get("session_id") or "").strip()
    responsible_id_raw = (request.args.get("responsible_id") or "").strip()
    product_id_raw = (request.args.get("product_id") or "").strip()
    from_date_raw = (request.args.get("from_date") or "").strip()
    to_date_raw = (request.args.get("to_date") or "").strip()

    report_payload, error_response = _build_sales_report_payload(
        period=period,
        session_id_raw=session_id_raw,
        responsible_id_raw=responsible_id_raw,
        product_id_raw=product_id_raw,
        from_date_raw=from_date_raw,
        to_date_raw=to_date_raw,
    )
    if error_response is not None:
        return error_response

    return jsonify(report_payload)


def _build_sales_report_payload(
    period: str,
    session_id_raw: str,
    responsible_id_raw: str,
    product_id_raw: str,
    from_date_raw: str,
    to_date_raw: str,
) -> tuple[dict[str, Any] | None, tuple[Response, int] | None]:

    db_path = current_app.config["DB_PATH"]
    where_clauses = ["o.order_status = 'PAID'"]
    params: list[Any] = []

    parsed_session_id: int | None = None
    parsed_responsible_id: int | None = None
    parsed_product_id: int | None = None
    parsed_from_date: str | None = None
    parsed_to_date: str | None = None
    from_date_start: datetime | None = None
    to_date_start: datetime | None = None

    if period in {"today", "week"}:
        now = datetime.now(UTC)
        start = datetime(now.year, now.month, now.day, tzinfo=UTC)
        if period == "week":
            start = start - timedelta(days=6)
        where_clauses.append("o.created_at >= ?")
        params.append(start.isoformat())

    if session_id_raw:
        try:
            parsed_session_id = int(session_id_raw)
        except ValueError:
            return None, (jsonify({"error": "session_id must be an integer."}), 400)
        where_clauses.append("o.session_id = ?")
        params.append(parsed_session_id)

    if responsible_id_raw:
        try:
            parsed_responsible_id = int(responsible_id_raw)
        except ValueError:
            return None, (jsonify({"error": "responsible_id must be an integer."}), 400)
        where_clauses.append("ps.opened_by = ?")
        params.append(parsed_responsible_id)

    if product_id_raw:
        try:
            parsed_product_id = int(product_id_raw)
        except ValueError:
            return None, (jsonify({"error": "product_id must be an integer."}), 400)
        where_clauses.append("EXISTS (SELECT 1 FROM order_items oi WHERE oi.order_id = o.id AND oi.product_id = ?)")
        params.append(parsed_product_id)

    if from_date_raw:
        try:
            from_date_start = datetime.strptime(from_date_raw, "%Y-%m-%d").replace(tzinfo=UTC)
        except ValueError:
            return None, (jsonify({"error": "from_date must be in YYYY-MM-DD format."}), 400)
        parsed_from_date = from_date_raw
        where_clauses.append("o.created_at >= ?")
        params.append(from_date_start.isoformat())

    if to_date_raw:
        try:
            to_date_start = datetime.strptime(to_date_raw, "%Y-%m-%d").replace(tzinfo=UTC)
        except ValueError:
            return None, (jsonify({"error": "to_date must be in YYYY-MM-DD format."}), 400)
        parsed_to_date = to_date_raw
        where_clauses.append("o.created_at < ?")
        params.append((to_date_start + timedelta(days=1)).isoformat())

    if from_date_start and to_date_start and from_date_start > to_date_start:
        return None, (jsonify({"error": "from_date cannot be later than to_date."}), 400)

    where_sql = " AND ".join(where_clauses)

    with get_connection(db_path) as connection:
        summary = connection.execute(
            f"""
            SELECT
                COALESCE(SUM(o.total_amount), 0) AS total_sales,
                COUNT(o.id) AS order_count,
                COALESCE(AVG(o.total_amount), 0) AS avg_order_value
            FROM orders o
            JOIN pos_sessions ps ON ps.id = o.session_id
            WHERE {where_sql};
            """,
            tuple(params),
        ).fetchone()

        product_rows = connection.execute(
            f"""
            SELECT
                p.id AS product_id,
                p.name AS product_name,
                COALESCE(SUM(oi.quantity), 0) AS quantity_sold,
                COALESCE(SUM(oi.line_subtotal + oi.line_tax), 0) AS revenue
            FROM orders o
            JOIN pos_sessions ps ON ps.id = o.session_id
            JOIN order_items oi ON oi.order_id = o.id
            JOIN products p ON p.id = oi.product_id
            WHERE {where_sql}
            GROUP BY p.id, p.name
            ORDER BY revenue DESC
            LIMIT 10;
            """,
            tuple(params),
        ).fetchall()

        payment_rows = connection.execute(
            f"""
            SELECT
                pm.method_type,
                pm.name,
                COUNT(p.id) AS transactions,
                COALESCE(SUM(p.amount), 0) AS total
            FROM orders o
            JOIN pos_sessions ps ON ps.id = o.session_id
            LEFT JOIN payments p ON p.order_id = o.id AND p.payment_status = 'CONFIRMED'
            LEFT JOIN payment_methods pm ON pm.id = p.payment_method_id
            WHERE {where_sql}
            GROUP BY pm.method_type, pm.name
            ORDER BY total DESC;
            """,
            tuple(params),
        ).fetchall()

    return (
        {
            "filters": {
                "period": period,
                "session_id": parsed_session_id,
                "responsible_id": parsed_responsible_id,
                "product_id": parsed_product_id,
                "from_date": parsed_from_date,
                "to_date": parsed_to_date,
            },
            "summary": {
                "total_sales": round(float(summary["total_sales"]), 2),
                "order_count": int(summary["order_count"]),
                "avg_order_value": round(float(summary["avg_order_value"]), 2),
            },
            "by_product": [
                {
                    "product_id": row["product_id"],
                    "product_name": row["product_name"],
                    "quantity_sold": int(row["quantity_sold"]),
                    "revenue": round(float(row["revenue"]), 2),
                }
                for row in product_rows
            ],
            "by_payment_method": [
                {
                    "method_type": row["method_type"] or "UNKNOWN",
                    "name": row["name"] or "Unknown",
                    "transactions": int(row["transactions"]),
                    "total": round(float(row["total"]), 2),
                }
                for row in payment_rows
                if row["transactions"] > 0
            ],
        },
        None,
    )


@pos_bp.get("/reports/sales/export")
@auth_required
def export_sales_report():
    export_format = (request.args.get("format") or "csv").strip().lower()
    if export_format not in {"csv", "xlsx", "pdf"}:
        return jsonify({"error": "format must be csv, xlsx or pdf."}), 400

    period = (request.args.get("period") or "all").strip().lower()
    session_id_raw = (request.args.get("session_id") or "").strip()
    responsible_id_raw = (request.args.get("responsible_id") or "").strip()
    product_id_raw = (request.args.get("product_id") or "").strip()
    from_date_raw = (request.args.get("from_date") or "").strip()
    to_date_raw = (request.args.get("to_date") or "").strip()

    report_payload, error_response = _build_sales_report_payload(
        period=period,
        session_id_raw=session_id_raw,
        responsible_id_raw=responsible_id_raw,
        product_id_raw=product_id_raw,
        from_date_raw=from_date_raw,
        to_date_raw=to_date_raw,
    )
    if error_response is not None:
        return error_response

    timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")

    if export_format == "csv":
        rows = []
        rows.append(["Odoo POS Cafe - Sales Report"])
        rows.append(["Generated At", datetime.now(UTC).isoformat()])
        rows.append(["Period", report_payload["filters"]["period"]])
        rows.append(["From Date", report_payload["filters"]["from_date"] or "ALL"])
        rows.append(["To Date", report_payload["filters"]["to_date"] or "ALL"])
        rows.append([])
        rows.append(["Summary"])
        rows.append(["Total Sales", report_payload["summary"]["total_sales"]])
        rows.append(["Order Count", report_payload["summary"]["order_count"]])
        rows.append(["Average Order Value", report_payload["summary"]["avg_order_value"]])
        rows.append([])
        rows.append(["Top Products"])
        rows.append(["Product Id", "Product Name", "Quantity Sold", "Revenue"])
        for row in report_payload["by_product"]:
            rows.append(
                [
                    row["product_id"],
                    row["product_name"],
                    row["quantity_sold"],
                    row["revenue"],
                ]
            )
        rows.append([])
        rows.append(["Payment Mix"])
        rows.append(["Method Type", "Name", "Transactions", "Total"])
        for row in report_payload["by_payment_method"]:
            rows.append(
                [
                    row["method_type"],
                    row["name"],
                    row["transactions"],
                    row["total"],
                ]
            )

        from io import StringIO

        sio = StringIO()
        writer = csv.writer(sio)
        writer.writerows(rows)
        content = sio.getvalue()

        return Response(
            content,
            mimetype="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=sales_report_{timestamp}.csv"
            },
        )

    if export_format == "pdf":
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
        except ModuleNotFoundError:
            return jsonify({"error": "PDF export requires reportlab. Install: pip install reportlab"}), 500

        output = BytesIO()
        pdf = canvas.Canvas(output, pagesize=A4)
        page_width, page_height = A4
        y_pos = page_height - 42

        def write_line(text: str, bold: bool = False, spacing: float = 14.0) -> None:
            nonlocal y_pos
            if y_pos <= 42:
                pdf.showPage()
                y_pos = page_height - 42
            pdf.setFont("Helvetica-Bold" if bold else "Helvetica", 10)
            pdf.drawString(36, y_pos, text)
            y_pos -= spacing

        write_line("Odoo POS Cafe - Sales Report", bold=True, spacing=20)
        write_line(f"Generated At: {datetime.now(UTC).isoformat()}")
        write_line(f"Period: {report_payload['filters']['period']}")
        write_line(f"From Date: {report_payload['filters']['from_date'] or 'ALL'}")
        write_line(f"To Date: {report_payload['filters']['to_date'] or 'ALL'}")
        write_line(
            "Filters -> "
            f"Session: {report_payload['filters']['session_id'] or 'ALL'}, "
            f"Responsible: {report_payload['filters']['responsible_id'] or 'ALL'}, "
            f"Product: {report_payload['filters']['product_id'] or 'ALL'}"
        )
        y_pos -= 8

        write_line("Summary", bold=True)
        write_line(f"Total Sales: {report_payload['summary']['total_sales']}")
        write_line(f"Order Count: {report_payload['summary']['order_count']}")
        write_line(f"Average Order Value: {report_payload['summary']['avg_order_value']}")
        y_pos -= 8

        write_line("Top Products", bold=True)
        if report_payload["by_product"]:
            for row in report_payload["by_product"]:
                write_line(
                    f"#{row['product_id']} {row['product_name']} | Qty: {row['quantity_sold']} | Revenue: {row['revenue']}"
                )
        else:
            write_line("No product sales for selected filters.")
        y_pos -= 8

        write_line("Payment Mix", bold=True)
        if report_payload["by_payment_method"]:
            for row in report_payload["by_payment_method"]:
                write_line(
                    f"{row['name']} ({row['method_type']}) | Txn: {row['transactions']} | Total: {row['total']}"
                )
        else:
            write_line("No payment data for selected filters.")

        pdf.save()
        return Response(
            output.getvalue(),
            mimetype="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=sales_report_{timestamp}.pdf"
            },
        )

    try:
        from openpyxl import Workbook
    except ModuleNotFoundError:
        return jsonify({"error": "XLSX export requires openpyxl. Install: pip install openpyxl"}), 500

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["Generated At", datetime.now(UTC).isoformat()])
    summary_sheet.append(["Period", report_payload["filters"]["period"]])
    summary_sheet.append(["From Date", report_payload["filters"]["from_date"] or "ALL"])
    summary_sheet.append(["To Date", report_payload["filters"]["to_date"] or "ALL"])
    summary_sheet.append(["Total Sales", report_payload["summary"]["total_sales"]])
    summary_sheet.append(["Order Count", report_payload["summary"]["order_count"]])
    summary_sheet.append(["Average Order Value", report_payload["summary"]["avg_order_value"]])

    product_sheet = workbook.create_sheet("Top Products")
    product_sheet.append(["Product Id", "Product Name", "Quantity Sold", "Revenue"])
    for row in report_payload["by_product"]:
        product_sheet.append(
            [
                row["product_id"],
                row["product_name"],
                row["quantity_sold"],
                row["revenue"],
            ]
        )

    payment_sheet = workbook.create_sheet("Payment Mix")
    payment_sheet.append(["Method Type", "Name", "Transactions", "Total"])
    for row in report_payload["by_payment_method"]:
        payment_sheet.append(
            [
                row["method_type"],
                row["name"],
                row["transactions"],
                row["total"],
            ]
        )

    output = BytesIO()
    workbook.save(output)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=sales_report_{timestamp}.xlsx"
        },
    )
